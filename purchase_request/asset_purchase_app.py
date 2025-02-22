import tkinter as tk
import pandas as pd

class OutputManager:
    def __init__(self, asset_manager):
        self.asset_manager = asset_manager
        self.clipboard_formatter = PurchaseRequestFormatter()
        self.excel_formatter = PurchaseRequestExcelStyleGuide()
    
    def prepare_clipboard_data(self):
        """클립보드에 복사할 데이터 준비"""
        # 현재 선택된 또는 가장 최근 asset을 가져옴
        assets = self.asset_manager.assets.get_all_assets()
        if not assets:
            raise ValueError("No assets available")
        return self.clipboard_formatter.format_data(assets[0]) #TODO : check this
    
    def prepare_excel_data(self):
        """엑셀 출력용 데이터 준비"""
        assets = self.asset_manager.assets.get_all_assets()
        if not assets:
            raise ValueError("No assets available")
            
        data = []
        for asset in assets:
            data.append({
                "구분": asset.category,
                "품명": asset.item_name,
                "수량": asset.quantity,
                "단가": asset.unit_price,
                "금액": asset.total_price
            })
        return pd.DataFrame(data)
    
class ClipboardService:
    def __init__(self, root):
        self.root = root
        
    def copy_text(self, text):
        """입력된 정보를 바탕으로 텍스트를 clipboard에 복사합니다."""
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        msgbox.showinfo("클립보드 복사 완료", "텍스트가 클립보드에 복사되었습니다.")
        

class DocumentFormatter:
    """문서 형식을 정의하는 기본 클래스"""
    def format_data(self, data):
        raise NotImplementedError
    
class PurchaseRequestFormatter(DocumentFormatter):
    """구매요청서 형식을 정의하는 클래스"""
    def format_data(self, asset):
        """Asset 객체를 구매요청서 형식의 텍스트로 변환"""
        return f"""
상기의 건에 대하여 아래와 같이 구매하고자 하오니 검토 후 재가 바랍니다.

                                                - 아 래 -

1. 목 적 : {asset.purpose}(계정명 : 공기구비품) 


2. 내 용
{asset.content}


3. 구입내역 


4. 구입업체 : (주) {asset.vendor}


5. 결제조건 : 현금결제


6. 납품 기한 및 방법 : 발주 후 2주 이내 


7. 첨부 : 견적서 1부
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class PurchaseRequestExcelStyleGuide:
    """엑셀 스타일 가이드를 정의하는 클래스"""
    def __init__(self):
        self.header_font = Font(color="FFFFFF")
        self.header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        self.total_font = Font(color="FFFFFF")
        self.total_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        self.total_alignment = Alignment(horizontal='center')
        self.currency_format = '₩#,##0'

import os

import datetime

class ExcelService:
    def __init__(self, style_guide):
        self.style = style_guide
    
    def save(self, df):
        """DataFrame을 엑셀 파일로 저장"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = '(VAT 별도)'
        
        # DataFrame을 엑셀에 쓰기
        for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False), start=2):
            ws.append(row)
            if r_idx == 2:  # Header styling
                for cell in ws[2]:
                    cell.font = self.style.header_font
                    cell.fill = self.style.header_fill

        # 합계 행 추가 및 스타일링
        total_row = ['합계', '', df['수량'].sum(), '-', df['금액'].sum()]
        ws.append(total_row)
        for cell in ws[len(df) + 3]:
            cell.font = self.style.total_font
            cell.fill = self.style.total_fill
            cell.alignment = self.style.total_alignment
        
        ws.cell(row=len(df) + 3, column=5).number_format = self.style.currency_format

        # 숫자 형식 지정
        for row in range(2, len(df) + 3):
            ws.cell(row=row, column=4).number_format = self.style.currency_format
            ws.cell(row=row, column=5).number_format = self.style.currency_format
        
        # 각 열의 너비 자동 조정
        def get_formatted_length(value, number_format):
            """포맷팅된 값의 길이를 반환"""
            if isinstance(value, (int, float)):
                # currency_format이 '₩#,##0'라고 가정
                formatted = format(value, ',')
                return len('₩' + formatted)
            return len(str(value))

        # 각 열의 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if column in ['D', 'E']:  # 4,5번째 열
                        length = get_formatted_length(cell.value, self.style.currency_format)
                    else:
                        length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
            adjusted_width = (max_length + 2)  # 여백 추가
            ws.column_dimensions[column].width = adjusted_width

        # 파일 저장
        now = datetime.datetime.now()
        file_name = f"구매요청서_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(file_name)
        os.startfile(file_name)
        msgbox.showinfo("엑셀 열기 완료", "엑셀 파일이 열렸습니다.")


from tkinter import ttk

class View:
   def __init__(self, master):
        self.master = master
        self.master.title("Asset Purchase App")
        
        # data access
        self.datas = Assets()
        self.asset_manager = AssetManager(self.datas)
        self.output_manager = OutputManager(self.asset_manager)
        
        # services
        self.clipboard = ClipboardService(master)
        self.excel = ExcelService(PurchaseRequestExcelStyleGuide())
        
        # Treeview setup
        self.setup_treeview()
    
   def setup_treeview(self):
        """Initialize treeview with modified headers"""
        self.input_headers = ["품명", "수량", "단가"]
        self.tree_headers = ["ID"] + self.input_headers + ["구분", "금액"]
        self.tree = ttk.Treeview(self.master, columns=self.tree_headers, show="headings")
        
        # Set column headings
        for header in self.tree_headers:
            self.tree.heading(header, text=header)
        
        # Column widths
        self.tree.column("ID", width=50)
        self.tree.column("품명", width=150)
        self.tree.column("수량", width=70)
        self.tree.column("단가", width=70)
        self.tree.column("구분", width=70)
        self.tree.column("금액", width=70)
        
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        
   def render_view(self):
        # Common information section (top)
        tk.Label(self.master, text="").grid(row=0, pady=10)
        
        # Fixed inputs frame
        fixed_frame = tk.LabelFrame(self.master, text="공통 정보", padx=5, pady=5)
        fixed_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        
        tk.Label(fixed_frame, text="목적\t:").grid(row=0, column=0, sticky="e", pady=5)
        self.purpose_entry = tk.Entry(fixed_frame, width=50)
        self.purpose_entry.grid(row=0, column=1, sticky="w")
        
        tk.Label(fixed_frame, text="내용\t:").grid(row=1, column=0, sticky="e", pady=5)
        self.inner_content_entry = tk.Text(fixed_frame, width=50, height=5)
        self.inner_content_entry.grid(row=1, column=1, sticky="w")
        
        tk.Label(fixed_frame, text="구입업체\t:").grid(row=2, column=0, sticky="e", pady=5)
        self.selling_company_entry = tk.Entry(fixed_frame, width=50)
        self.selling_company_entry.grid(row=2, column=1, sticky="w")
        
        # Item details frame
        items_frame = tk.LabelFrame(self.master, text="품목 정보", padx=5, pady=5)
        items_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        
        tk.Label(items_frame, text="품명\t:").grid(row=0, column=0, sticky="e", pady=5)
        self.item_name_entry = tk.Entry(items_frame, width=50)
        self.item_name_entry.grid(row=0, column=1, sticky="w")
        
        tk.Label(items_frame, text="수량\t:").grid(row=1, column=0, sticky="e", pady=5)
        self.quantity_entry = tk.Entry(items_frame, width=50)
        self.quantity_entry.grid(row=1, column=1, sticky="w")
        
        tk.Label(items_frame, text="단가\t:").grid(row=2, column=0, sticky="e", pady=5)
        self.unit_price_entry = tk.Entry(items_frame, width=50)
        self.unit_price_entry.grid(row=2, column=1, sticky="w")
        
        # Buttons for item manipulation
        manipulate_buttons_frame = tk.Frame(self.master)
        manipulate_buttons_frame.grid(row=3, column=0, columnspan=2, pady=5)
        
        tk.Button(manipulate_buttons_frame, text="품목 추가", command=self.add_row).pack(side="left", padx=5)
        tk.Button(manipulate_buttons_frame, text="품목 수정", command=self.modify_row).pack(side="left", padx=5)
        tk.Button(manipulate_buttons_frame, text="품목 삭제", command=self.delete_row).pack(side="left", padx=5)
        
        # Treeview
        self.tree.grid(row=4, column=0, columnspan=2, padx=10, pady=5, sticky="nsew")
        
        # Output buttons
        result_buttons_frame = tk.Frame(self.master)
        result_buttons_frame.grid(row=5, column=0, columnspan=2, pady=10)
        
        tk.Button(result_buttons_frame, text="텍스트 클립보드에 복사", 
                 command=lambda: self.copy_text_to_clipboard()).pack(side="left", padx=5)
        tk.Button(result_buttons_frame, text="차트 엑셀로 열기", 
                 command=lambda: self.save_excel()).pack(side="left", padx=5)
    
   def clear_item_inputs(self):
        """Clear only item-related input fields"""
        self.item_name_entry.delete(0, tk.END)
        self.quantity_entry.delete(0, tk.END)
        self.unit_price_entry.delete(0, tk.END)
        

   def on_tree_select(self, event):
        """Handle treeview selection event"""
        selected_items = self.tree.selection()
        if not selected_items:
            return
            
        # Get the selected item's values
        selected_values = self.tree.item(selected_items[0])['values']
        
        # Clear only item inputs
        self.clear_item_inputs()
        
        # Fill only item-related fields
        self.item_name_entry.insert(0, selected_values[1])  # 품명
        self.quantity_entry.insert(0, selected_values[2])   # 수량
        self.unit_price_entry.insert(0, selected_values[3]) # 단가
    
   def add_row(self):
        """Add a new item row"""
        data = Asset.builder() \
            .purpose(self.purpose_entry.get()) \
            .content(self.inner_content_entry.get("1.0", tk.END)) \
            .vendor(self.selling_company_entry.get()) \
            .item_name(self.item_name_entry.get()) \
            .quantity(int(self.quantity_entry.get().replace(",", ""))) \
            .unit_price(int(self.unit_price_entry.get().replace(",", ""))) \
            .build()
            
        self.asset_manager.add_asset(data)
        
        # Only show item-related data in treeview
        tree_data = [data.id, data.item_name, data.quantity, 
                    data.unit_price, data.category, data.total_price]
        self.tree.insert("", "end", values=tree_data)
        
        self.clear_item_inputs()

        
   def modify_row(self):
        """Modify selected item row"""
        selected_item = self.tree.selection()
        if not selected_item:
            tk.messagebox.showwarning("Warning", "수정할 품목을 선택해주세요")
            return

        selected_values = self.tree.item(selected_item[0])["values"]
        asset_id = selected_values[0]
        
        modified_asset = Asset.builder() \
            .id(asset_id) \
            .purpose(self.purpose_entry.get()) \
            .content(self.inner_content_entry.get("1.0", tk.END)) \
            .vendor(self.selling_company_entry.get()) \
            .item_name(self.item_name_entry.get()) \
            .quantity(int(self.quantity_entry.get().replace(",", ""))) \
            .unit_price(int(self.unit_price_entry.get().replace(",", ""))) \
            .build()
        
        self.asset_manager.modify_asset(asset_id, modified_asset)
        
        # Update treeview with only item-related data
        tree_data = [modified_asset.id, modified_asset.item_name, 
                    modified_asset.quantity, modified_asset.unit_price, 
                    modified_asset.category, modified_asset.total_price]
        self.tree.item(selected_item, values=tree_data)
        
        self.clear_item_inputs()
    
   def delete_row(self):
        """Delete selected item row"""
        selected_items = self.tree.selection()
        if not selected_items:
            tk.messagebox.showwarning("Warning", "삭제할 품목을 선택해주세요")
            return
            
        if tk.messagebox.askyesno("Delete", "선택한 품목을 삭제하시겠습니까?"):
            for selected_record in selected_items:
                asset_id = self.tree.item(selected_record)['values'][0]
                self.asset_manager.remove_asset(asset_id)
                self.tree.delete(selected_record)
            
            self.clear_item_inputs()

    
   def copy_text_to_clipboard(self):
        try:
            text = self.output_manager.prepare_clipboard_data()
            self.clipboard.copy_text(text)
        except Exception as e:
            msgbox.showerror("Error", str(e))
        
   def save_excel(self):
        try:
            data = self.output_manager.prepare_excel_data()
            self.excel.save(data)
        except Exception as e:
            msgbox.showerror("Error", str(e))
        
from tkinter import messagebox as msgbox

class Asset:
    def __init__(self):
        self.id = None
        self.purpose = None
        self.content = None
        self.vendor = None
        self.item_name = None
        self.quantity = 1
        self.unit_price = None
        self.category = "신규 구매"
        self.total_price = None
    
    def __eq__(self, other):
        if not isinstance(other, Asset):
            return False
        return self.id == other.id
    
    def __hash__(self):
        return hash(self.id)
    
    @staticmethod
    def builder():
        return Asset.Builder()

    class Builder:
        def __init__(self):
            self.asset = Asset()

        def purpose(self, purpose):
            self.asset.purpose = purpose
            return self

        def content(self, content):
            self.asset.content = content
            return self

        def vendor(self, vendor):
            self.asset.vendor = vendor
            return self

        def item_name(self, item_name):
            self.asset.item_name = item_name
            return self

        def quantity(self, quantity):
            self.asset.quantity = int(quantity)
            return self

        def unit_price(self, unit_price):
            self.asset.unit_price = int(unit_price)
            self.asset.total_price = int(self.asset.quantity) * int(self.asset.unit_price)
            return self

        def category(self, category):
            self.asset.category = category
            return self
        
        def id(self, id):
            self.asset.id = id
            return self

        def build(self):
            return self.asset

class Assets:
    def __init__(self):
        self.assets = []
        self._next_id = 1
    
    def generate_id(self):
        id = self._next_id
        self._next_id += 1
        return id
    
    def add_asset(self, asset):
        if isinstance(asset, Asset):
            if asset.id is None:
                asset.id = self.generate_id()
            self.assets.append(asset)
        else:
            raise TypeError("Only Asset objects can be added")

    def remove_asset(self, asset_id):
        self.assets = [asset for asset in self.assets if asset.id != asset_id]
        
    def get_asset_by_id(self, asset_id):
        for asset in self.assets:
            if asset.id == asset_id:
                return asset
        return None
    
    def get_all_assets(self):
        return self.assets

    def get_asset_by_index(self, index):
        return self.assets[index]

    def get_total_value(self):
        return sum(asset.total_price for asset in self.assets if asset.total_price is not None)

    def get_assets_by_category(self, category):
        return [asset for asset in self.assets if asset.category == category]

    def get_assets_by_vendor(self, vendor):
        return [asset for asset in self.assets if asset.vendor == vendor]

    def __len__(self):
        return len(self.assets)

    def __iter__(self):
        return iter(self.assets)

class AssetManager:
    def __init__(self, assets):
        self.assets = assets
    
    def add_asset(self, asset):
        self.assets.add_asset(asset)
    
    def remove_asset(self, asset_id):
        self.assets.remove_asset(asset_id)
    
    def modify_asset(self, asset_id, new_asset):
        # Remove old asset and add the modified one
        self.remove_asset(asset_id)
        self.add_asset(new_asset)
    
class AssetPurchaseApp:
    def __init__(self, master):
        
        self.master = master
        
        self.view = View(self.master)
        self.view.render_view()