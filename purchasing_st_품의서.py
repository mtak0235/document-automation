import tkinter as tk
import tkinter.messagebox as msgbox
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
import os
import subprocess
import datetime

def copy_text_to_clipboard():
    """입력된 정보를 바탕으로 텍스트를 clipboard에 복사합니다."""
    purpose_text = purpose_entry.get()
    inner_content_text = inner_content_entry.get("1.0", tk.END)
    selling_company = selling_company_entry.get()
    content = f"""
상기의 건에 대하여 아래와 같이 구매하고자 하오니 검토 후 재가 바랍니다.

- 아 래 -
1. 목 적 : {purpose_text}(계정명 : 공기구비품) 

2. 내 용
{inner_content_text}

3. 구입내역 (자세한 내용은 엑셀 파일 참조)

4. 구입업체 : (주) {selling_company}

5. 결제조건 : 현금결제

6. 납품 기한 및 방법 : 발주 후 2주 이내 

7. 첨부 : 견적서 1부
"""

    # Clipboard에 복사
    root.clipboard_clear()
    root.clipboard_append(content)
    root.update()

    msgbox.showinfo("클립보드 복사 완료", "텍스트가 클립보드에 복사되었습니다.")

def open_excel_with_chart_selected():
    """표 데이터를 엑셀에 저장하고, 해당 영역을 선택한 채로 엑셀을 엽니다."""
    # 표 데이터 생성
    data = []
    for i in range(len(category_entries)):
        category = category_entries[i].get()
        item_name = item_name_entries[i].get()
        quantity = int(quantity_entries[i].get())
        unit_price = int(unit_price_entries[i].get().replace(',', ''))  # 쉼표 제거
        amount = int(amount_entries[i].get().replace(',', ''))  # 쉼표 제거
        data.append([category, item_name, quantity, unit_price, amount])

    # pandas DataFrame으로 표 생성
    df = pd.DataFrame(data, columns=["구분", "품명", "수량", "단가", "금액"])
    
    # 합계 계산
    total_amount = df['금액'].astype(float).sum()
    total_quantity = df['수량'].astype(float).sum()
    
    # 합계 행 추가
    total_row = pd.DataFrame([['합계', '', total_quantity, '-', total_amount]], columns=["구분", "품명", "수량", "단가", "금액"])
    df = pd.concat([df, total_row], ignore_index=True)

    # 엑셀 파일 생성 및 데이터 쓰기
    wb = Workbook()
    ws = wb.active

    # DataFrame을 엑셀에 쓰기
    for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False)):
        ws.append(row)
        if r_idx == 0:  # 첫 번째 행 (헤더) 스타일 변경
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF")  # 흰색 글자
                cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # 진한 파랑 배경
        if row[0] == '합계':  # 합계 행 스타일 변경
            for cell in ws[len(data) + 2]:
                cell.font = Font(color="FFFFFF")  # 흰색 글자
                cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # 진한 파랑 배경
                cell.alignment = Alignment(horizontal='center')  # 가운데 정렬
    
    # VAT 별도 텍스트 추가
    ws.cell(row=1, column=6).value = "(VAT 별도)"
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.column_dimensions["F"].width = 12
    
    # 숫자 형식 지정
    num_rows = len(data) + 2
    for row in range(2, num_rows + 1):
        ws.cell(row=row, column=4).number_format = '₩#,##0'
        ws.cell(row=row, column=5).number_format = '₩#,##0'

    # "품명" 열 너비 조정
    max_length = 0
    for row in range(2, num_rows + 1):
        cell_value = str(ws.cell(row=row, column=2).value)
        max_length = max(max_length, len(cell_value))
    ws.column_dimensions["B"].width = max_length + 5  # 가장 긴 품명 + 여유 공간

    # 열 너비 조정
    for col in ["D", "E"]:  # "단가"와 "금액" 열
        ws.column_dimensions[col].width = 15  # 열 너비 15로 설정
    
    # 합계 행의 "구분"과 "품명" 셀 병합
    ws.merge_cells(start_row=len(data) + 2, start_column=1, end_row=len(data) + 2, end_column=2)

    # 파일 이름을 현재 시간 기준으로 생성
    now = datetime.datetime.now()
    file_name = f"구매요청서_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_file = file_name

    # 임시 엑셀 파일로 저장
    wb.save(excel_file)

    # 엑셀 파일 열기 (기본 엑셀 프로그램으로)
    os.startfile(excel_file)

    msgbox.showinfo("엑셀 열기 완료", "엑셀 파일이 열렸습니다.")

def add_row():
    """표에 새로운 행을 추가합니다."""
    global category_entries, item_name_entries, quantity_entries, unit_price_entries, amount_entries
    
    category_entry = tk.Entry(table_frame)
    item_name_entry = tk.Entry(table_frame)
    quantity_entry = tk.Entry(table_frame)
    unit_price_entry = tk.Entry(table_frame)
    amount_entry = tk.Entry(table_frame)
    
    category_entry.grid(row=len(category_entries) + 1, column=0)
    item_name_entry.grid(row=len(item_name_entries) + 1, column=1)
    quantity_entry.grid(row=len(quantity_entries) + 1, column=2)
    unit_price_entry = tk.Entry(table_frame)
    amount_entry = tk.Entry(table_frame)
    
    category_entry.grid(row=len(category_entries) + 1, column=0)
    item_name_entry.grid(row=len(item_name_entries) + 1, column=1)
    quantity_entry.grid(row=len(quantity_entries) + 1, column=2)
    unit_price_entry.grid(row=len(unit_price_entries) + 1, column=3)
    amount_entry.grid(row=len(amount_entries) + 1, column=4)
    
    category_entries.append(category_entry)
    item_name_entries.append(item_name_entry)
    quantity_entries.append(quantity_entry)
    unit_price_entries.append(unit_price_entry)
    amount_entries.append(amount_entry)

# Tkinter 윈도우 생성
root = tk.Tk()
root.title("구매 요청서 생성기")

# 입력 필드
tk.Label(root, text="목적:").grid(row=0, column=0, sticky="e")
purpose_entry = tk.Entry(root, width=50)
purpose_entry.grid(row=0, column=1, sticky="w")

tk.Label(root, text="내용:").grid(row=1, column=0, sticky="e")
inner_content_entry = tk.Text(root, width=50, height=5)
inner_content_entry.grid(row=1, column=1, sticky="w")

tk.Label(root, text="구입업체:").grid(row=3, column=0, sticky="e")
selling_company_entry = tk.Entry(root, width=50)
selling_company_entry.grid(row=3, column=1, sticky="w")

# 표 입력을 위한 프레임
table_frame = tk.Frame(root)
table_frame.grid(row=4, column=0, columnspan=2)

# 표 헤더
tk.Label(table_frame, text="구분").grid(row=0, column=0)
tk.Label(table_frame, text="품명").grid(row=0, column=1)
tk.Label(table_frame, text="수량").grid(row=0, column=2)
tk.Label(table_frame, text="단가").grid(row=0, column=3)
tk.Label(table_frame, text="금액").grid(row=0, column=4)

# 표 입력 필드 리스트 (초기 행 1개)
category_entries = []
item_name_entries = []
quantity_entries = []
unit_price_entries = []
amount_entries = []

add_row()

# 행 추가 버튼
add_row_button = tk.Button(root, text="행 추가", command=add_row)
add_row_button.grid(row=5, column=0, columnspan=2)

# 텍스트 클립보드 복사 버튼
copy_text_button = tk.Button(root, text="텍스트 클립보드에 복사", command=copy_text_to_clipboard)
copy_text_button.grid(row=6, column=0, columnspan=2)

# 차트 엑셀로 저장 및 열기 버튼
copy_chart_button = tk.Button(root, text="차트 엑셀로 열기", command=open_excel_with_chart_selected)
copy_chart_button.grid(row=7, column=0, columnspan=2)

root.mainloop()
