import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import copy
import pyperclip  # 클립보드 사용을 위한 pyperclip import

class AssetDisposalApp:
    def __init__(self, master):
        self.master = master
        master.title("폐기대상List")

        # 데이터 저장 리스트
        self.data = []

        # 폰트 및 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        self.currency_format = '#,##0 ₩'
        self.skyblue_fill = PatternFill(start_color="A9B5DF", end_color="A9B5DF", fill_type="solid")
        self.deepblue_fill = PatternFill(start_color="2D336B", end_color="2D336B", fill_type="solid")
        self.light_purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        # 레이블 및 입력 필드 생성 (간격 조정)
        label_padx = 1
        entry_padx = 1
        label_pady = 1
        entry_pady = 1
        column_width = 15

        self.category_label = tk.Label(master, text="구분:")
        self.category_label.grid(row=0, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.category_entry = tk.Entry(master, width=column_width)
        self.category_entry.grid(row=0, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.asset_number_label = tk.Label(master, text="자산번호:")
        self.asset_number_label.grid(row=1, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.asset_number_entry = tk.Entry(master, width=column_width)
        self.asset_number_entry.grid(row=1, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.asset_name_label = tk.Label(master, text="고정자산명칭:")
        self.asset_name_label.grid(row=2, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.asset_name_entry = tk.Entry(master, width=column_width)
        self.asset_name_entry.grid(row=2, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.management_number_label = tk.Label(master, text="전산관리번호:")
        self.management_number_label.grid(row=3, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.management_number_entry = tk.Entry(master, width=column_width)
        self.management_number_entry.grid(row=3, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.acquisition_date_label = tk.Label(master, text="취득일 (YYYY-MM-DD):")
        self.acquisition_date_label.grid(row=4, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.acquisition_date_entry = tk.Entry(master, width=column_width)
        self.acquisition_date_entry.grid(row=4, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.acquisition_cost_label = tk.Label(master, text="취득원가:")
        self.acquisition_cost_label.grid(row=5, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.acquisition_cost_entry = tk.Entry(master, width=column_width)
        self.acquisition_cost_entry.grid(row=5, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.accumulated_depreciation_label = tk.Label(master, text="상각누계액:")
        self.accumulated_depreciation_label.grid(row=6, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.accumulated_depreciation_entry = tk.Entry(master, width=column_width)
        self.accumulated_depreciation_entry.grid(row=6, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.disposal_value_label = tk.Label(master, text="처분자산가액:")
        self.disposal_value_label.grid(row=7, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.disposal_value_entry = tk.Entry(master, width=column_width)
        self.disposal_value_entry.grid(row=7, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        self.disposal_loss_label = tk.Label(master, text="처분손익:")
        self.disposal_loss_label.grid(row=8, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.disposal_loss_entry = tk.Entry(master, width=column_width)
        self.disposal_loss_entry.grid(row=8, column=1, padx=entry_padx, pady=entry_pady, sticky="w")

        # 자산상태 선택 (토글 버튼)
        self.status_label = tk.Label(master, text="자산상태:")
        self.status_label.grid(row=9, column=0, padx=label_padx, pady=label_pady, sticky="w")
        self.status = tk.BooleanVar()
        self.status_checkbutton = tk.Checkbutton(master, text="전체 폐기", variable=self.status, onvalue=True, offvalue=False)
        self.status_checkbutton.grid(row=9, column=1, padx=entry_padx, pady=entry_pady, sticky="w")
        self.status_checkbutton.select() # Default value: 전체 폐기...
        self.save_button = tk.Button(master, text="저장", command=self.save_data)
        self.save_button.grid(row=10, column=0, padx=5, pady=5, sticky="w")

        self.export_button = tk.Button(master, text="엑셀 Export", command=self.export_to_excel)
        self.export_button.grid(row=10, column=1, padx=5, pady=5, sticky="w")

        # "Context 복사" 버튼 추가
        self.copy_context_button = tk.Button(master, text="Context 복사", command=self.copy_context_to_clipboard)
        self.copy_context_button.grid(row=10, column=2, padx=5, pady=5, sticky="w")  # Export 버튼 옆에 배치

        # 데이터 테이블 (Treeview 활용)
        self.tree = ttk.Treeview(master, columns=("구분", "자산번호", "고정자산명칭", "전산관리번호", "관리부서", "취득일", "자산상태", "취득원가", "상각누계액", "상각잔액", "처분자산가액", "처분손익"), show="headings")
        self.tree.heading("#1", text="구분")
        self.tree.heading("#2", text="자산번호")
        self.tree.heading("#3", text="고정자산명칭")
        self.tree.heading("#4", text="전산관리번호")
        self.tree.heading("#5", text="관리부서")
        self.tree.heading("#6", text="취득일")
        self.tree.heading("#7", text="자산상태")
        self.tree.heading("#8", text="취득원가")
        self.tree.heading("#9", text="상각누계액")
        self.tree.heading("#10", text="상각잔액")
        self.tree.heading("#11", text="처분자산가액")
        self.tree.heading("#12", text="처분손익")
        
        # 각 열의 너비 설정
        self.tree.column("구분", width=70)
        self.tree.column("자산번호", width=70)
        self.tree.column("고정자산명칭", width=250)
        self.tree.column("전산관리번호", width=180)
        self.tree.column("관리부서", width=80)
        self.tree.column("취득일", width=80)
        self.tree.column("자산상태", width=70)
        self.tree.column("취득원가", width=80)
        self.tree.column("상각누계액", width=80)
        self.tree.column("상각잔액", width=80)
        self.tree.column("처분자산가액", width=90)
        self.tree.column("처분손익", width=80)

        self.tree.grid(row=11, column=0, columnspan=2, padx=5, pady=5, sticky="w")

    def calculate_depreciation_balance(self, acquisition_cost, accumulated_depreciation):
        """상각잔액 계산"""
        try:
            cost = float(acquisition_cost)
            depreciation = float(accumulated_depreciation)
            return cost - depreciation
        except ValueError:
            return "오류"

    def save_data(self):
        """입력 데이터 저장 및 테이블에 추가"""
        category = self.category_entry.get()
        asset_number = self.asset_number_entry.get()
        asset_name = self.asset_name_entry.get()
        management_number = self.management_number_entry.get()
        acquisition_date = self.acquisition_date_entry.get()
        status = "전체 폐기" if self.status.get() else "부분 폐기"
        acquisition_cost = self.acquisition_cost_entry.get()
        accumulated_depreciation = self.accumulated_depreciation_entry.get()
        disposal_value = self.disposal_value_entry.get()
        disposal_loss = self.disposal_loss_entry.get()

        # 상각잔액 계산
        depreciation_balance = self.calculate_depreciation_balance(acquisition_cost, accumulated_depreciation)

        # 데이터 유효성 검사 (예: 필수 필드 확인)
        if not all([category, asset_number, asset_name, management_number, acquisition_date, status, acquisition_cost, accumulated_depreciation, disposal_value, disposal_loss]):
            messagebox.showerror("오류", "모든 필드를 입력해주세요.")
            return

        # 데이터 테이블에 추가
        self.tree.insert("", "end", values=(category, asset_number, asset_name, management_number, "정보기획팀", acquisition_date, status, acquisition_cost, accumulated_depreciation, depreciation_balance, disposal_value, disposal_loss))

        # 데이터 리스트에 저장
        self.data.append([category, asset_number, asset_name, management_number, "정보기획팀", acquisition_date, status, acquisition_cost, accumulated_depreciation, depreciation_balance, disposal_value, disposal_loss])

        # 입력 필드 초기화
        self.category_entry.delete(0, tk.END)
        self.asset_number_entry.delete(0, tk.END)
        self.asset_name_entry.delete(0, tk.END)
        self.management_number_entry.delete(0, tk.END)
        self.acquisition_date_entry.delete(0, tk.END)
        self.acquisition_cost_entry.delete(0, tk.END)
        self.accumulated_depreciation_entry.delete(0, tk.END)
        self.disposal_value_entry.delete(0, tk.END)
        self.disposal_loss_entry.delete(0, tk.END)

    def export_to_excel(self):
        """데이터를 엑셀 파일로 Export"""
        if not self.data:
            messagebox.showinfo("알림", "저장된 데이터가 없습니다.")
            return

        try:
            workbook = openpyxl.Workbook()

            # 첫 번째 시트: 폐기상세YYYYMM
            current_year = datetime.now().year
            current_month = datetime.now().month
            sheet_name = f"폐기상세{current_year}{str(current_month).zfill(2)}"
            sheet = workbook.active
            sheet.title = sheet_name

            # 헤더 작성
            header = ["구분", "자산번호", "고정자산명칭", "전산관리번호", "관리부서", "취득일", "자산상태", "취득원가", "상각누계액", "상각잔액", "처분자산가액", "처분손익"]
            for col, column_title in enumerate(header, 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center')
                cell.fill = self.deepblue_fill

            # 데이터 작성
            row_num = 2
            category_start_row = 2
            current_category = None
            category_counts = {}
            category_totals = {}
            grand_total = {
                'count': 0,
                'acquisition_cost': 0,
                'accumulated_depreciation': 0,
                'depreciation_balance': 0,
                'disposal_value': 0,
                'disposal_loss': 0
            }

            for data_row in self.data:
                category = data_row[0]
                asset_number = data_row[1]  # 자산번호
                grand_total['count'] += 1

                # 숫자 데이터 추출 및 합계 계산
                try:
                    acquisition_cost = float(data_row[7])
                    accumulated_depreciation = float(data_row[8])
                    depreciation_balance = float(data_row[9])
                    disposal_value = float(data_row[10])
                    disposal_loss = float(data_row[11])

                    # 총계 누적
                    grand_total['acquisition_cost'] += acquisition_cost
                    grand_total['accumulated_depreciation'] += accumulated_depreciation
                    grand_total['depreciation_balance'] += depreciation_balance
                    grand_total['disposal_value'] += disposal_value
                    grand_total['disposal_loss'] += disposal_loss

                    # 카테고리별 합계 초기화 및 누적
                    if category not in category_totals:
                        category_totals[category] = {
                            'acquisition_cost': 0,
                            'accumulated_depreciation': 0,
                            'depreciation_balance': 0,
                            'disposal_value': 0,
                            'disposal_loss': 0,
                            'asset_numbers': []  # 자산번호 리스트
                        }
                    category_totals[category]['acquisition_cost'] += acquisition_cost
                    category_totals[category]['accumulated_depreciation'] += accumulated_depreciation
                    category_totals[category]['depreciation_balance'] += depreciation_balance
                    category_totals[category]['disposal_value'] += disposal_value
                    category_totals[category]['disposal_loss'] += disposal_loss
                    category_totals[category]['asset_numbers'].append(asset_number) # 자산번호 추가
                except (ValueError, TypeError):
                    pass

                # 카테고리 변경 시 시작 행 업데이트
                if category != current_category:
                    if current_category is not None:
                        # 소계 행 추가
                        for col in range(1, 13):
                            cell = sheet.cell(row=row_num, column=col)
                            cell.border = Border(left=Side(style='thin'),
                                                right=Side(style='thin'),
                                                top=Side(style='thin'),
                                                bottom=Side(style='thin'))
                            cell.fill = self.skyblue_fill

                        # "소계" 텍스트와 병합
                        sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                        cell = sheet.cell(row=row_num, column=1)
                        cell.value = "소계"
                        cell.alignment = Alignment(horizontal='center')

                        # 개수 표시
                        record_count = category_counts[current_category]
                        cell = sheet.cell(row=row_num, column=3)
                        cell.value = f"{record_count}"
                        cell.alignment = Alignment(horizontal='center')

                        # 금액 합계 표시
                        if current_category in category_totals:
                            # 상각잔액
                            cell = sheet.cell(row=row_num, column=10)
                            cell.value = category_totals[current_category]['depreciation_balance']
                            cell.number_format = self.currency_format
                            cell.alignment = Alignment(horizontal='right')
                        row_num += 1

                    category_start_row = row_num
                    current_category = category
                    category_counts[category] = 0

                # 데이터 행 작성
                for col, cell_value in enumerate(data_row, 1):
                    cell = sheet.cell(row=row_num, column=col)
                    cell.value = cell_value
                    cell.border = self.border

                    # 금액 관련 열은 통화 형식 적용 및 오른쪽 정렬
                    if col in [8, 9, 10, 11, 12]:
                        try:
                            cell.value = float(cell_value)
                            cell.number_format = self.currency_format
                            cell.alignment = Alignment(horizontal='right')
                        except ValueError:
                            pass

                category_counts[category] += 1
                row_num += 1

            # 마지막 카테고리 소계 처리
            if current_category is not None:
                # 소계 행 추가
                for col in range(1, 13):
                    cell = sheet.cell(row=row_num, column=col)
                    cell.border = Border(left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin'))
                    cell.fill = self.skyblue_fill
                # "소계" 텍스트와 병합
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                cell = sheet.cell(row=row_num, column=1)
                cell.value = "소계"
                cell.alignment = Alignment(horizontal='center')

                # 개수 표시
                record_count = category_counts[current_category]
                cell = sheet.cell(row=row_num, column=3)
                cell.value = f"{record_count}"
                cell.alignment = Alignment(horizontal='center')

                # 금액 합계 표시
                if current_category in category_totals:
                    # 상각잔액
                    cell = sheet.cell(row=row_num, column=10)
                    cell.value = category_totals[current_category]['depreciation_balance']
                    cell.number_format = self.currency_format
                    cell.alignment = Alignment(horizontal='right')
                row_num += 1

            # 총계 계산 및 작성
            total_row_num = row_num + 1

            # 총계 행 스타일 적용
            for col in range(1, 13):
                cell = sheet.cell(row=total_row_num, column=col)
                cell.border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                cell.fill = self.light_purple_fill

            # "총계" 텍스트와 병합
            sheet.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=2)
            cell = sheet.cell(row=total_row_num, column=1)
            cell.value = "총계"
            cell.alignment = Alignment(horizontal='center')

            # 총 개수 표시
            cell = sheet.cell(row=total_row_num, column=3)
            cell.value = grand_total['count']
            cell.alignment = Alignment(horizontal='center')

            # 총 금액 표시
            cell = sheet.cell(row=total_row_num, column=8)
            cell.value = grand_total['acquisition_cost']
            cell.number_format = self.currency_format
            cell.alignment = Alignment(horizontal='right')

            cell = sheet.cell(row=total_row_num, column=9)
            cell.value = grand_total['accumulated_depreciation']
            cell.number_format = self.currency_format
            cell.alignment = Alignment(horizontal='right')

            # 총 상각잔액
            cell = sheet.cell(row=total_row_num, column=10)
            cell.value = grand_total['depreciation_balance']
            cell.number_format = self.currency_format
            cell.alignment = Alignment(horizontal='right')

            cell = sheet.cell(row=total_row_num, column=11)
            cell.value = grand_total['disposal_value']
            cell.number_format = self.currency_format
            cell.alignment = Alignment(horizontal='right')

            cell = sheet.cell(row=total_row_num, column=12)
            cell.value = grand_total['disposal_loss']
            cell.number_format = self.currency_format
            cell.alignment = Alignment(horizontal='right')

            # 열 너비 조정
            for col in range(1, 13):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 15

            # 엑셀 파일 저장
            filename = f"폐기대상_{current_year}{str(current_month).zfill(2)}.xlsx"
            workbook.save(filename)
            messagebox.showinfo("알림", f"{filename} 파일로 저장되었습니다.")

        except Exception as e:
            messagebox.showerror("오류", f"엑셀 Export 중 오류가 발생했습니다: {e}")

    def copy_context_to_clipboard(self):
        """Context 내용을 클립보드에 복사"""
        context = f"""

상기의 건에 대하여 아래와 같이 폐기 및 고정자산 정리를 하고자 하오니 검토 후 재가 바랍니다.
 

                                                                - 아       래 -
 

1. 목    적 : 불용 전산장비 폐기 및 고정자산 정리

2. 내    용 :

 1) 불용 전산장비 폐기 대상                                                                



 2) 불용 전산장비 내역 및 폐기 사유 : 노후 및 파손으로 인한 사용불가 (수리불가)

 
 
 3) 폐기 방법 : 폐기 품목 중 사용 가능한 부품은 분리 활용 후 폐기\n\t폐기 승인 품목에 대해서는 재무팀에서 고정자산 정리

3. 폐기일시 : 품의 후 2주 이내

4. 수거업체 : (주)신도네트만승

5. 폐기비용 : 무상 수거

 

6. 첨부 : 폐기대상LIST 1부, 폐기대상 사진 1부
"""
        try:
            pyperclip.copy(context)
            messagebox.showinfo("알림", "Context가 클립보드에 복사되었습니다.")
        except pyperclip.PyperclipException as e:
            messagebox.showerror("오류", f"클립보드 복사 중 오류가 발생했습니다: {e}\nPyperclip이 설치되어 있는지 확인해주세요.")

root = tk.Tk()
app = AssetDisposalApp(root)
root.mainloop()
